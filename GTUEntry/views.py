from django.shortcuts import render, get_object_or_404,redirect
from .models import *
from faculty.models import *
from .forms import GTUExamDataForm, GTUExamCPForm,InvoiceForm,all_InvoiceForm,ExamSelectForm
from datetime import date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from authenticate.decorators import role_required
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import F
from num2words import num2words
# views.py
# # GTUEntry/views.py
from django.http import HttpResponse
from weasyprint import HTML
from django.template.loader import get_template
from django.template.loader import render_to_string
import io
import zipfile
#exel file
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.db.models import Count



@login_required()
@role_required(['admin','gtu_cordinator'])
def select_pdf(request):
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        if form.is_valid():
            faculty = form.cleaned_data['select_faculty']
            exam = form.cleaned_data['select_exam'] 
            fac = list(gtu_th_exam_data.objects.filter(faculty_id=faculty,exam_id=exam))
            # Create a map of duty types and their payable amounts from exam_name
            #role_amount = fac.duty_type_id.duty_type

            duty_field_map = {
                'GTU Coordinator': 'gtu_co',
                'Center Incharge': 'center_incharge',
                'Senior Supervisor': 'sr_sup',
                'Junior Supervisor': 'jr_sup',
                'Reliever Supervisor': 'st_sup',
                'Numerator Supervisor': 'num_sup',
                'Peon': 'peon',
                'Stationary Peon': 'st_peon',
                'Sweeper': 'sweeper',
                'Paper Checking': 'page_amount',
            }

            # Add payable amounts to each duty object
            total_amount = 0
            for duty in fac:
                duty_name = duty.duty_type_id.duty  # e.g., "Junior Supervisor"
                exam_field = duty_field_map.get(duty_name)

                if exam_field:
                    payable_amount = getattr(exam, exam_field, 0)
                else:
                    payable_amount = 0

                # Attach amount temporarily to each duty object
                duty.payable_amount = payable_amount
                total_amount += payable_amount
            rows = []
            for i in range(21):  # 1 to 21 on left, 22 to 42 on right
                left = fac[i] if i < len(fac) else None
                right = fac[i + 21] if (i + 21) < len(fac) else None

                row = {
                    "no_left": i + 1,
                    "left": left,
                    "no_right": i + 22,
                    "right": right,
                }
                rows.append(row)

            amount_in_words = num2words(total_amount, to='cardinal').title()

            context = {
                "faculty": faculty,
                "exam": exam,
                "rows": rows,
                "fac": fac,
                "total_amount": total_amount,
                "amount_in_words": amount_in_words,
           }
             # Render HTML to PDF


            return render(request, 'pdf_genrate.html', context)
    else:
        form = InvoiceForm()
    return render(request, 'print_pdf_selection.html', {'form': form})
# def exam_data_view(request):
#     form = GTUExamDataForm()
#     return render(request, 'add_th_exam_entry.html', {'form': form})
@login_required()
@role_required(['admin','gtu_cordinator'])
def download_pdf(request, faculty_id, exam_id):
    faculty_obj = faculty.objects.get(id=faculty_id)
    exam_obj = exam_name.objects.get(id=exam_id)
    fac = list(gtu_th_exam_data.objects.filter(faculty_id=faculty_obj))

    duty_field_map = {
        'GTU Coordinator': 'gtu_co',
        'Center Incharge': 'center_incharge',
        'Senior Supervisor': 'sr_sup',
        'Junior Supervisor': 'jr_sup',
        'Reliever Supervisor': 'st_sup',
        'Numerator Supervisor': 'num_sup',
        'Peon': 'peon',
        'Stationary Peon': 'st_peon',
        'Sweeper': 'sweeper',
        'Paper Checking': 'page_amount',
    }

    total_amount = 0
    for duty in fac:
        duty_name = duty.duty_type_id.duty
        exam_field = duty_field_map.get(duty_name)
        payable_amount = getattr(exam_obj, exam_field, 0) if exam_field else 0
        duty.payable_amount = payable_amount
        total_amount += payable_amount

    rows = []
    for i in range(21):
        left = fac[i] if i < len(fac) else None
        right = fac[i + 21] if (i + 21) < len(fac) else None
        rows.append({
            "no_left": i + 1, "left": left,
            "no_right": i + 22, "right": right,
        })

    amount_in_words = num2words(total_amount, lang='en').capitalize() + " only"

    context = {
        "faculty": faculty_obj,
        "exam": exam_obj,
        "rows": rows,
        "fac": fac,
        "total_amount": total_amount,
        "amount_in_words": amount_in_words,
    }

    html_string = get_template('pdfformat.html').render(context)
    pdf = HTML(string=html_string).write_pdf()
    file_name = f"{faculty_obj.short_name}_{exam_obj}_invoice.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{file_name}"'
    return response


@login_required()
@role_required(['admin', 'gtu_cordinator'])
def allpdf(request, exam_id):
    exam_obj = exam_name.objects.get(id=exam_id)
    faculty_list = faculty.objects.filter(
        id__in=gtu_th_exam_data.objects.filter(exam_id=exam_obj).values_list('faculty_id', flat=True).distinct()
    )

    # ZIP file buffer
    zip_buffer = io.BytesIO()
    zip_file = zipfile.ZipFile(zip_buffer, 'w')

    for fac in faculty_list:
        # Get duty entries for this faculty in this exam
        fac_duties = list(gtu_th_exam_data.objects.filter(faculty_id=fac, exam_id=exam_obj))

        total_amount = 0
        for duty in fac_duties:
            duty_field_map = {
                'GTU Coordinator': 'gtu_co',
                'Center Incharge': 'center_incharge',
                'Senior Supervisor': 'sr_sup',
                'Junior Supervisor': 'jr_sup',
                'Reliever Supervisor': 'st_sup',
                'Numerator Supervisor': 'num_sup',
                'Peon': 'peon',
                'Stationary Peon': 'st_peon',
                'Sweeper': 'sweeper',
                'Paper Checking': 'page_amount',
            }
            duty_name = duty.duty_type_id.duty
            exam_field = duty_field_map.get(duty_name)
            duty.payable_amount = getattr(exam_obj, exam_field, 0) if exam_field else 0
            total_amount += duty.payable_amount

        # Split duties into left and right columns
        rows = []
        for i in range(21):
            left = fac_duties[i] if i < len(fac_duties) else None
            right = fac_duties[i+21] if (i+21) < len(fac_duties) else None
            rows.append({"no_left": i + 1, "left": left, "no_right": i + 22, "right": right})

        # Render to PDF
        context = {
            "faculty": fac,
            "exam": exam_obj,
            "rows": rows,
            "fac": fac_duties,
            "total_amount": total_amount,
            "amount_in_words": num2words(total_amount, lang='en').capitalize() + " only"
        }

        html_string = get_template('pdfformat.html').render(context)
        pdf = HTML(string=html_string).write_pdf()

        file_name = f"{fac.short_name}_Duty_Report.pdf"
        zip_file.writestr(file_name, pdf)

    zip_file.close()
    zip_buffer.seek(0)

    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="All_Faculty_Reports.zip"'
    
    return response          
    


@login_required()
@role_required(['admin','gtu_cordinator'])
def exam_data_view(request):
    # Get active faculty to display initially
    active_faculty = faculty.objects.filter(is_active=True).order_by('full_name')
    initial_data = {'date': date.today()}

    if request.method == 'POST':
        form = GTUExamDataForm(request.POST)
        if form.is_valid():
            form.save()
            # Add success message or redirect
    else:
        form = GTUExamDataForm(initial=initial_data)

    return render(request, 'add_th_exam_entry.html', {
        'form': form,
        'active_faculty': active_faculty  # Pass active faculty to template
    })


@login_required()
@role_required(['admin','gtu_cordinator'])
def load_times(request):
    session_id = request.GET.get('session_id')
    times = time.objects.filter(session_id=session_id).order_by('time')
    return render(request, 'times_dropdown.html', {'times': times})

def select_exam_for_pdf(request):
    if request.method == 'POST':
        form = all_InvoiceForm(request.POST)
        if form.is_valid():
            selected_exam = form.cleaned_data['select_exam']
            return redirect('generate_all_pdfs', exam_id=selected_exam.id)  # name of your URL pattern
    else:
        form = all_InvoiceForm()
    return render(request, 'select_exam_pdf.html', {'form': form})

# def load_faculty(request):
#     search = request.GET.get('search', '')
#     if search:
#         faculty_list = faculty.objects.filter(
#             is_active=True,
#             full_name__icontains=search
#         ).order_by('full_name')
#     else:
#         faculty_list = faculty.objects.filter(is_active=True).order_by('full_name')
#
#     return render(request, 'faculty_dropdown.html', {'faculty_list': faculty_list})

@login_required()
@role_required(['admin','gtu_cordinator'])
def load_faculty(request):
    search_term = request.GET.get('search', '')
    faculty_list = faculty.objects.filter(is_active=True)

    if search_term:
        faculty_list = faculty_list.filter(
            models.Q(full_name__icontains=search_term) |
            models.Q(short_name__icontains=search_term)
        )

    return render(request, 'faculty_dropdown.html', {
        'faculty_list': faculty_list.order_by('full_name')
    })
@login_required
@role_required(['admin', 'gtu_cordinator'])
def select_exam_for_excel(request):
    if request.method == 'POST':
        form = ExamSelectForm(request.POST)
        if form.is_valid():
            selected_exam = form.cleaned_data['exam']
            return redirect('generate_excel', exam_id=selected_exam.id)
    else:
        form = ExamSelectForm()
    return render(request, 'select_exam_excel.html', {'form': form})


@login_required
@role_required(['admin', 'gtu_cordinator'])
def generate_excel(request, exam_id):
    exam_obj = exam_name.objects.get(id=exam_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "GTU CP BILL DATA"

    # Title Header
    ws.merge_cells('A1:V1')
    ws['A1'] = "GTU CP BILL DATA (WINTER-2024)"
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = [
        "Sr.No", "Date", "Session", "Sem", "NoofStud", "No of Block", 
        "Center Incharrge", "GTU Cordinator", "Sr.Sup", "Jr.Sup",
        "St.Cum Relsup", "NumCum RelSup", "Peon", "St.Cum NoPeon",
        "Swp", "Total Pages", "Page Amount",
        "Center Incharrge", "GTU Cordinator", "Sr.Sup", "Jr.Sup",
        "St.Cum Relsup", "NumCum RelSup", "Peon", "St.Cum NoPeon",
        "Swp", "TOTAL",
    ]
    ws.append(headers)
    total_of_all_rows = 0

    grouped_data = gtu_th_CP_data.objects.filter(exam_id=exam_obj).values(
        'date', 'session_id', 'semester_id'
    ).distinct()

    for idx, group in enumerate(grouped_data, start=1):
        date = group['date']
        session_id = group['session_id']
        semester_id = group['semester_id']

        # Get actual related objects
        session_obj = session.objects.get(id=session_id)
        semester_obj = semester.objects.get(id=semester_id)

        # Duty counts per type
        summary = gtu_th_exam_data.objects.filter(
            exam_id=exam_obj,
            date=date,
            session_id=session_obj,
            semester_id=semester_obj
        ).values(
            'duty_type_id__duty'
        ).annotate(
            count=Count('id')
        )

        roles = {
            "Center Incharge": 0,
            "GTU Coordinator": 0,
            "Senior Supervisor": 0,
            "Junior Supervisor": 0,
            "Reliever Supervisor": 0,
            "Numbering Supervisor": 0,
            "Peon": 0,
            "Stationary Peon": 0,
            "Sweeper": 0
        }

        for item in summary:
            duty = item['duty_type_id__duty']
            if duty in roles:
                roles[duty] = item['count']

        # Fetch CP data for total pages and students
        get_a_ob = gtu_th_CP_data.objects.filter(
            exam_id=exam_obj,
            date=date,
            session_id=session_obj,
            semester_id=semester_obj
        )

        total_students = 0
        total_pages = 0  # Initialize total_pages before usage
        
        for obj in get_a_ob:
            total_pages += obj.total_pages or 0
            print(f"✅ Processing CP data ID: {obj.id}")
            for i in range(1, 13):
                students = getattr(obj, f"sub{i}_no_of_student", 0) or 0
                print(f"   sub{i}_no_of_student = {students}")
                total_students += students
        print(f"\n🎯 Total Students: {total_students}")

        # Calculate the page amount based on total pages
        page_amount = total_pages * exam_obj.page_amount if total_pages else 0

        # Calculate total amounts for each role by multiplying the count by the rate from `exam_name`
        center_incharge_amount = roles["Center Incharge"] * exam_obj.center_incharge
        gtu_co_amount = roles["GTU Coordinator"] * exam_obj.gtu_co
        sr_sup_amount = roles["Senior Supervisor"] * exam_obj.sr_sup
        jr_sup_amount = roles["Junior Supervisor"] * exam_obj.jr_sup
        st_sup_amount = roles["Reliever Supervisor"] * exam_obj.st_sup
        num_sup_amount = roles["Numbering Supervisor"] * exam_obj.num_sup
        peon_amount = roles["Peon"] * exam_obj.peon
        st_peon_amount = roles["Stationary Peon"] * exam_obj.st_peon
        sweeper_amount = roles["Sweeper"] * exam_obj.sweeper

        # Calculate total amount (sum of all role amounts and page amount)
        total_amount = (
            center_incharge_amount +
            gtu_co_amount +
            sr_sup_amount +
            jr_sup_amount +
            st_sup_amount +
            num_sup_amount +
            peon_amount +
            st_peon_amount +
            sweeper_amount +
            page_amount
        )
        total_of_all_rows += total_amount

        # Create the row with all required information
        row = [
            idx,
            date.strftime("%d-%m-%Y"),
            session_obj.session,
            semester_obj.sem,
            total_students,
            roles["Junior Supervisor"],  # Assuming this = No of Block
            roles["Center Incharge"],
            roles["GTU Coordinator"],
            roles["Senior Supervisor"],
            roles["Junior Supervisor"],
            roles["Reliever Supervisor"],
            roles["Numbering Supervisor"],
            roles["Peon"],
            roles["Stationary Peon"],
            roles["Sweeper"],
            total_pages,
            page_amount,
            center_incharge_amount,
            gtu_co_amount,
            sr_sup_amount,
            jr_sup_amount,
            st_sup_amount,
            num_sup_amount,
            peon_amount,
            st_peon_amount,
            sweeper_amount,
            total_amount,
        ]

        # Append the row to the worksheet
        ws.append(row)
    total_row = ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "Ov.Total", total_of_all_rows]
    ws.append(total_row)
    # Prepare the response for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    dynamic_filename = f"gtu_cp_bill_{exam_obj.exam}_{date.strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{dynamic_filename}"'
    wb.save(response)
    return response








# @login_required()
# @role_required(['admin','gtu_cordinator'])
# def exam_entry(request):
#     exam = gtu_th_exam_data.objects.all()  # Fetch all exams
#     return render(request,'view_exams.html',{'exam': exam})


# @login_required()
# @role_required(['admin', 'gtu_cordinator'])
# def exam_views(request):
#     exam = gtu_th_exam_data.objects.all()  # Fetch all exams
#     return render(request,'exam_edit_view.html',{'exam': exam})
@login_required()
@role_required(['admin', 'gtu_cordinator'])
def exam_views(request):
    exam_id = request.GET.get("exam_id")
    if exam_id:
        exam = gtu_th_exam_data.objects.filter(exam_id=exam_id)
    else:
        exam = gtu_th_exam_data.objects.all()

    exams_list = exam_name.objects.all().order_by('-id')

    # If it's an HTMX request, return only the partial HTML
    if request.headers.get('HX-Request'):
        html = render_to_string('_exam_table_partial.html', {'exam': exam})
        return HttpResponse(html)

    # Otherwise, render full page
    return render(request, 'exam_edit_view.html', {
        'exam': exam,
        'exams_list': exams_list,
    })


@login_required()
@role_required(['admin', 'gtu_cordinator'])
def edit_exam(request,exam_id):
    exam_obj = get_object_or_404(gtu_th_exam_data, pk=exam_id)
    if request.method == 'POST':
        form = GTUExamDataForm(request.POST, instance=exam_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "exam updated successfully!")
            return redirect('exam_views')
        # else:
        #     messages.error(request, "Please correct the errors below.")
    else:
        form = GTUExamDataForm(instance=exam_obj)
    return render(request,'edit_exam.html', {'form': form})


@login_required()
@role_required(['admin', 'gtu_cordinator'])
def add_gtu_exam_cp(request):
    initial_date = {'date': date.today()}
    if request.method == 'POST':
        form = GTUExamCPForm(request.POST)
        subject_data = {}
        grand_total = 0

        # Extract all subject data from POST manually
        for i in range(1, 13):
            code_key = f'sub_{i}_code'
            name_key = f'sub_{i}_name'
            students_key = f'sub{i}_no_of_student'
            pages_key = f'sub{i}_no_of_pages'

            code = request.POST.get(code_key, '').strip()
            name = request.POST.get(name_key, '').strip()

            try:
                students = int(request.POST.get(students_key, 0))
            except ValueError:
                students = 0
            try:
                pages = int(request.POST.get(pages_key, 0))
            except ValueError:
                pages = 0

            subject_data[code_key] = code
            subject_data[name_key] = name
            subject_data[students_key] = students
            subject_data[pages_key] = pages

            grand_total += students * pages

        if form.is_valid():
            instance = form.save(commit=False)

            # Apply manual fields to the instance
            for key, value in subject_data.items():
                setattr(instance, key, value)

            instance.total_pages = grand_total
            instance.save()
            messages.success(request, "Data saved successfully.")
            return redirect('add_gtu_exam_cp')

        else:
            # Form is invalid, return all subject data to the template
            messages.error(request, "There was an error submitting the form. Please correct the issues below.")
            return render(request, 'exam_cp_data_form.html', {
                'form': form,
                'subject_range': range(1, 13),
                'subject_data': subject_data,  # optional, not used in template
            })

    else:
        form = GTUExamCPForm(initial=initial_date)
        return render(request, 'exam_cp_data_form.html', {
            'form': form,
            'subject_range': range(1, 13),
        })

@login_required()
@role_required(['admin', 'gtu_cordinator'])
def cp_exam_views(request):
    exam_id = request.GET.get("exam_id")
    if exam_id:
        cp_data = gtu_th_CP_data.objects.filter(exam_id=exam_id)
    else:
        cp_data = gtu_th_CP_data.objects.all()

    exams_list = exam_name.objects.all().order_by('-id')

    if request.headers.get('HX-Request'):
        html = render_to_string('_cp_table_partial.html', {'cp_data': cp_data})
        return HttpResponse(html)

    return render(request, 'cp_exam_view.html', {
        'cp_data': cp_data,
        'exams_list': exams_list,
    })


@login_required()
@role_required(['admin', 'gtu_cordinator'])
def edit_cp_exam(request, cp_id):
    cp_obj = get_object_or_404(gtu_th_CP_data, pk=cp_id)
    if request.method == 'POST':
        form = GTUExamCPForm(request.POST, instance=cp_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "CP exam updated successfully!")
            return redirect('cp_exam_views')
    else:
        form = GTUExamCPForm(instance=cp_obj)
    return render(request, 'edit_cp_exam.html', {'form': form})




@login_required()
@role_required(['admin', 'gtu_cordinator'])
def pdf_genrate(request):
    active_faculty = faculty.objects.filter(is_active=True).order_by('full_name')
    return render(request,'pdf_genrate.html',{'form':active_faculty})